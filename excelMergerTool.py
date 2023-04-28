import base64
import io
from openpyxl import load_workbook
import openpyxl as xl
from geteData import get_data
import base64
from openpyxl.styles import PatternFill, Side, Border


def excelMergerTool(base64String, OCRs, fileName):
    newWorkbook = xl.Workbook()
    # default_sheet = newWorkbook['Sheet']
    del newWorkbook["Sheet"]

    for i, val in enumerate(base64String):
        new_sheet = newWorkbook.create_sheet(OCRs[i])

        bytes_data = base64.b64decode(val)
        bytes_io = io.BytesIO(bytes_data)

        oldWorkbook = load_workbook(bytes_io)

        oldWorksheet = oldWorkbook.active

        mcOldWorksheet = oldWorksheet.max_column
        mrOldWorksheet = oldWorksheet.max_row

        for y in range(mcOldWorksheet):
            for x in range(mrOldWorksheet):
                new_sheet.cell(row=x + 1, column=y + 1).value = str(
                    oldWorksheet.cell(row=x + 1, column=y + 1).value
                )
                new_fill = PatternFill(
                    start_color=oldWorksheet.cell(
                        row=x + 1, column=y + 1
                    ).fill.start_color,
                    end_color=oldWorksheet.cell(row=x + 1, column=y + 1).fill.end_color,
                    fill_type=oldWorksheet.cell(row=x + 1, column=y + 1).fill.fill_type,
                )
                new_sheet.cell(row=x + 1, column=y + 1).fill = new_fill

    newWorkbook.save(f"{fileName}.xlsx")
