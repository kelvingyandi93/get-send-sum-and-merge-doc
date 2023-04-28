import base64
import io
from openpyxl import load_workbook
import openpyxl as xl
from geteData import get_data
import base64
from excelMergerTool import excelMergerTool
from openpyxl.styles import PatternFill


def excel_Processor(dateFrom, dateTo):
    # update_error_status("No Error")

    OCRs = ["OCR_STNK", "OCR_KTP", "OCR_NPWP", "OCR_BPKB", "OCR_KK"]

    myBase64 = []
    myError = []
    sum = []
    totalSum = 0
    isReady = False
    dataStatus = "Not ready"

    for OCR in OCRs:
        try:
            myBase64.append(get_data(dateFrom, dateTo, OCR))
        except Exception as e:
            myError.append(e)
            print(e)

    if len(myError) == 0:
        # print(myBase64)
        for i, val in enumerate(myBase64):
            bytes_data = base64.b64decode(val)
            bytes_io = io.BytesIO(bytes_data)

            workbook = load_workbook(bytes_io)

            worksheet = workbook.active
            tempSum = 0
            for row in worksheet.iter_rows(min_row=2, min_col=10, max_col=10):
                for cell in row:
                    value = int(cell.value)
                    if value == -1:
                        value = value * -1
                        # print(value)
                        tempSum += value
                    else:
                        continue
            # print(tempSum)
            sum.append(tempSum)
            totalSum += tempSum

        # print(len(sum))

        newWorkbook = xl.Workbook()
        newWorksheet = newWorkbook.active
        newWorksheet.cell(row=1, column=1).value = "Tipe OCR"
        newWorksheet.cell(row=1, column=2).value = "SUM"

        newWorksheet.cell(row=1, column=1).fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        newWorksheet.cell(row=1, column=2).fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        for i in range(len(OCRs)):
            newWorksheet.cell(row=i + 2, column=1).value = OCRs[i]
            newWorksheet.cell(row=i + 2, column=2).value = sum[i]

        newWorksheet.cell(row=len(OCRs) + 2, column=1).value = "Summary"
        newWorksheet.cell(row=len(OCRs) + 2, column=2).value = totalSum

        newWorksheet.cell(row=len(OCRs) + 2, column=1).fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        newWorksheet.cell(row=len(OCRs) + 2, column=2).fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        newWorkbook.save("example.xlsx")
        isReady = True

    if isReady == True:
        with open("example.xlsx", "rb") as file:
            myBase64.insert(0, base64.b64encode(file.read()).decode("utf-8"))

        excelMergerTool(
            myBase64,
            ["Summary", "OCR_STNK", "OCR_KTP", "OCR_NPWP", "OCR_BPKB", "OCR_KK"],
            f"data penggunaan {dateFrom} / {dateTo}",
        )

        dataStatus = "Ready"

    return sum, myError, dataStatus
